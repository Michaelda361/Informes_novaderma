from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
from datetime import datetime
import os
import base64
from io import BytesIO
from weasyprint import HTML, CSS
import tempfile

# Importaciones para ReportLab (alternativa más compatible)
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Importaciones para pdfkit (alternativa con wkhtmltopdf)
try:
    import pdfkit
    PDFKIT_AVAILABLE = True
except ImportError:
    PDFKIT_AVAILABLE = False

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Usar /tmp en producción (Render) o carpetas locales en desarrollo
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', 'uploads')
OUTPUT_FOLDER = os.environ.get('OUTPUT_FOLDER', 'output')

# Crear carpetas si no existen
for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# Variables globales para almacenar datos
evaluaciones_data = []

def get_logo_base64():
    """Convierte el logo a base64 para embebido en PDF"""
    logo_path = os.path.join('static', 'logo.png')
    try:
        if os.path.exists(logo_path):
            with open(logo_path, 'rb') as f:
                logo_data = f.read()
                base64_data = base64.b64encode(logo_data).decode('utf-8')
                print(f"Logo cargado exitosamente. Tamaño: {len(logo_data)} bytes")
                return base64_data
        else:
            print(f"Logo no encontrado en: {logo_path}")
            return ''
    except Exception as e:
        print(f"Error al cargar logo: {e}")
        return ''

def calcular_rendimiento(promedio):
    """Calcula la clasificación de rendimiento basada en el promedio"""
    if promedio >= 4.5:
        return "Sobresaliente"
    elif promedio >= 3.5:
        return "Satisfactorio"
    elif promedio >= 2.5:
        return "Aceptable"
    elif promedio >= 1.5:
        return "No Satisfactorio"
    else:
        return "Deficiente"

def normalizar_texto(texto):
    """Normaliza texto para búsqueda (sin acentos, mayúsculas, espacios)"""
    if not texto:
        return ""
    
    # Convertir a string y mayúsculas
    texto = str(texto).upper().strip()
    
    # Reemplazar acentos
    reemplazos = {
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'Ñ': 'N', 'Ü': 'U'
    }
    
    for acento, sin_acento in reemplazos.items():
        texto = texto.replace(acento, sin_acento)
    
    return texto

def encontrar_columna(headers, posibles_nombres, buscar_parcial=True):
    """Encuentra una columna por nombre, con múltiples opciones y búsqueda parcial"""
    headers_norm = [normalizar_texto(h) for h in headers]
    
    for nombre in posibles_nombres:
        nombre_norm = normalizar_texto(nombre)
        
        # Búsqueda exacta primero
        if nombre_norm in headers_norm:
            return headers_norm.index(nombre_norm)
        
        # Búsqueda parcial si está habilitada
        if buscar_parcial:
            for i, header in enumerate(headers_norm):
                if nombre_norm in header or header in nombre_norm:
                    return i
    
    return None

def detectar_tipo_evaluacion(headers):
    """Detecta el tipo de evaluación basado en el número de columnas y contenido"""
    num_cols = len(headers)
    
    print(f"Tipo de evaluación detectado: ", end="")
    
    if num_cols <= 51:
        print("OPERATIVO")
        return "OPERATIVO"
    elif num_cols <= 53:
        print("ADMINISTRATIVA") 
        return "ADMINISTRATIVA"
    elif num_cols <= 56:
        # Distinguir entre COMERCIAL y DIRECTIVOS por contenido
        headers_text = ' '.join(str(h).upper() for h in headers)
        if 'VENTAS' in headers_text or 'CUOTAS' in headers_text or 'MEDICO' in headers_text:
            print("COMERCIAL")
            return "COMERCIAL"
        else:
            print("DIRECTIVOS")
            return "DIRECTIVOS"
    else:
        print("DIRECTIVOS")
        return "DIRECTIVOS"

def extraer_calificaciones_por_categoria(headers, row, tipo_evaluacion):
    """Extrae calificaciones específicas según el tipo de evaluación"""
    calificaciones = {}
    
    def get_value(col_idx):
        if col_idx is not None and col_idx < len(row):
            val = row[col_idx]
            if val is not None and str(val).strip():
                try:
                    num_val = float(val)
                    # Solo aceptar valores en el rango válido de calificaciones (1-5)
                    if 1 <= num_val <= 5:
                        return num_val
                    else:
                        return 0
                except:
                    return 0
        return 0
    
    # Mapeo específico por tipo de evaluación
    if tipo_evaluacion == 'OPERATIVO':
        # Sección 3: Desempeño operativo
        operativo_cols = {
            'organizacion': ['Organiza las tareas a fin de cumplir con los tiempos establecidos'],
            'cumple_resultados': ['Cumple con los resultados esperados de su función'],
            'aportes_constructivos': ['Demuestra capacidad para apoyar y generar aportes constructivos al área'],
            'realiza_actividades': ['Realiza las actividades encomendadas según las instrucciones dadas']
        }
        for key, nombres in operativo_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
        
        # Sección 4: Compromiso y calidad
        compromiso_cols = {
            'cumple_politicas': ['Demuestra compromiso con el cumplimiento de los objetivos'],
            'conoce_calidad': ['Actúa en pro de los intereses de la empresa'],
            'propone_mejoras': ['Propone alternativas para mejorar el trabajo']
        }
        for key, nombres in compromiso_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
        
        # Sección 5: Comportamiento
        comportamiento_cols = {
            'relaciones': ['RELACIONES INTERPERSONALES', 'Mantiene relaciones de cordialidad'],
            'trabajo_equipo': ['TRABAJO EN EQUIPO', 'Apoya a los compañeros'],
            'actitud_servicio': ['ACTITUD DE SERVICIO', 'Se preocupa por satisfacer']
        }
        for key, nombres in comportamiento_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
    
    elif tipo_evaluacion == 'ADMINISTRATIVA':
        # Sección 3: Desempeño en el cargo
        admin_cols = {
            'organizacion': ['Conoce y aplica los procedimientos del área'],
            'cumple_resultados': ['Cumple con los resultados esperados de su función'],
            'aportes_constructivos': ['Demuestra capacidad para apoyar y generar aportes constructivos'],
            'realiza_actividades': ['Realiza las actividades encomendadas según las instrucciones'],
            'analisis': ['Demuestra capacidad para analizar y solucionar los problemas'],
            'informes': ['Presenta informes, cartas, etc., de manera oportuna'],
            'aplica_capacitacion': ['Aplica en su desempeño diario los conceptos vistos en capacitaciones'],
            'uso_equipos': ['Hace uso adecuado del equipo y demás elementos'],
            'entrega_tareas': ['Entrega los informes o tareas encomendadas']
        }
        for key, nombres in admin_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
        
        # Sección 4: Compromiso y calidad
        compromiso_cols = {
            'cumple_politicas': ['Demuestra compromiso con el cumplimiento de los objetivos'],
            'conoce_calidad': ['Actúa en pro de los intereses de la empresa'],
            'propone_mejoras': ['Propone alternativas para mejorar el trabajo']
        }
        for key, nombres in compromiso_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
        
        # Sección 5: Comportamiento
        comportamiento_cols = {
            'relaciones': ['RELACIONES INTERPERSONALES'],
            'trabajo_equipo': ['TRABAJO EN EQUIPO'],
            'actitud_servicio': ['ACTITUD DE SERVICIO']
        }
        for key, nombres in comportamiento_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
    
    elif tipo_evaluacion == 'COMERCIAL':
        # Calificaciones base
        comercial_cols = {
            'organizacion': ['ORGANIZA', 'ORGANIZACION', 'PLANIFICA'],
            'cumple_resultados': ['CUMPLE', 'RESULTADOS', 'OBJETIVOS'],
            'aplica_capacitacion': ['APLICA', 'CAPACITACION', 'ENTRENAMIENTO'],
            'uso_equipos': ['USO', 'EQUIPOS', 'RECURSOS'],
            'ventas': ['VENTAS', 'CUOTAS', 'PANEL MEDICO'],
            'clientes': ['CLIENTES', 'MEDICOS', 'ATENCION']
        }
        for key, nombres in comercial_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
        
        # Compromiso y calidad
        compromiso_cols = {
            'cumple_politicas': ['CUMPLE', 'POLITICAS', 'PROCEDIMIENTOS'],
            'conoce_calidad': ['CALIDAD', 'POLITICA DE CALIDAD'],
            'propone_mejoras': ['MEJORAS', 'IDEAS', 'ALTERNATIVAS']
        }
        for key, nombres in compromiso_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
        
        # Comportamiento
        comportamiento_cols = {
            'relaciones': ['RELACIONES', 'INTERPERSONALES', 'CORDIALIDAD'],
            'trabajo_equipo': ['EQUIPO', 'COLABORACION'],
            'actitud_servicio': ['SERVICIO', 'ACTITUD']
        }
        for key, nombres in comportamiento_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
    
    elif tipo_evaluacion == 'DIRECTIVOS':
        # Calificaciones específicas de directivos
        directivos_cols = {
            'organizacion': ['ORGANIZA', 'ORGANIZACION', 'PLANIFICA'],
            'cumple_resultados': ['CUMPLE', 'RESULTADOS', 'OBJETIVOS'],
            'aplica_capacitacion': ['APLICA', 'CAPACITACION', 'ENTRENAMIENTO'],
            'uso_equipos': ['USO', 'EQUIPOS', 'RECURSOS'],
            'liderazgo': ['LIDERAZGO', 'DIRECCION'],
            'gestion': ['GESTION', 'EFICIENTE'],
            'evaluacion_equipo': ['EVALUACION', 'ANALISIS', 'EQUIPO']
        }
        for key, nombres in directivos_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
        
        # Compromiso y calidad
        compromiso_cols = {
            'cumple_politicas': ['CUMPLE', 'POLITICAS', 'PROCEDIMIENTOS'],
            'conoce_calidad': ['CALIDAD', 'POLITICA DE CALIDAD'],
            'propone_mejoras': ['MEJORAS', 'IDEAS', 'ALTERNATIVAS']
        }
        for key, nombres in compromiso_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
        
        # Comportamiento
        comportamiento_cols = {
            'relaciones': ['RELACIONES', 'INTERPERSONALES', 'CORDIALIDAD'],
            'trabajo_equipo': ['EQUIPO', 'COLABORACION'],
            'actitud_servicio': ['SERVICIO', 'ACTITUD']
        }
        for key, nombres in comportamiento_cols.items():
            col_idx = encontrar_columna(headers, nombres, buscar_parcial=True)
            calificaciones[key] = get_value(col_idx)
    
    return calificaciones

def procesar_excel(file_path):
    """Procesa el archivo Excel y extrae las evaluaciones"""
    global evaluaciones_data
    evaluaciones_data = []
    
    try:
        # Cargar con data_only=True para obtener valores calculados de fórmulas
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Obtener headers
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(cell_value if cell_value else f"Col_{col}")
        
        # Detectar tipo de evaluación
        tipo_evaluacion = detectar_tipo_evaluacion(headers)
        
        # Definir columnas principales
        col_nombre = encontrar_columna(headers, ['NOMBRE', 'COLABORADOR'])
        col_cargo = encontrar_columna(headers, ['CARGO', 'PUESTO'])
        col_area = encontrar_columna(headers, ['AREA', 'PROCESO', 'DEPARTAMENTO'])
        col_jefe = encontrar_columna(headers, ['JEFE', 'SUPERVISOR', 'INMEDIATO'])
        col_fecha = encontrar_columna(headers, ['FECHA', 'EVALUACION'])
        
        # Buscar columna de periodo con mayor especificidad
        col_periodo = encontrar_columna(headers, ['PERIODO EVALUADO', 'PERÍODO EVALUADO'], buscar_parcial=False)
        if col_periodo is None:
            col_periodo = encontrar_columna(headers, ['PERIODO', 'PERÍODO'], buscar_parcial=True)
            # Verificar que no sea una columna de análisis
            if col_periodo is not None and col_periodo < len(headers):
                col_name = str(headers[col_periodo]).upper()
                if 'ANALISIS' in col_name or 'ANÁLISIS' in col_name:
                    col_periodo = None  # Ignorar esta columna
        
        # Buscar columna de promedio/porcentaje (más específica)
        col_promedio = encontrar_columna(headers, ['PROMEDIO', 'PORCENTAJE'], buscar_parcial=False)
        if col_promedio is None:
            # Búsqueda más amplia si no encuentra exacto
            col_promedio = encontrar_columna(headers, ['PROMEDIO', 'PORCENTAJE', '%'], buscar_parcial=True)
        
        # Buscar columnas de texto con mayor especificidad
        col_comentario = encontrar_columna(headers, ['COMENTARIOS DEL JEFE INMEDIATO', 'COMENTARIO DEL JEFE', 'FORTALEZAS Y DEBILIDADES'], buscar_parcial=True)
        col_aportes = encontrar_columna(headers, ['QUE APORTES HIZO USTED', 'QUE APORTES CONSIDERA', 'APORTES'], buscar_parcial=True)
        col_plan = encontrar_columna(headers, ['PLAN DE MEJORA PROPUESTO'], buscar_parcial=True)
        
        print(f"Columnas de texto encontradas - Comentario: {col_comentario}, Aportes: {col_aportes}, Plan: {col_plan}")
        if col_comentario is not None:
            print(f"  -> Comentario: '{headers[col_comentario]}'")
        if col_aportes is not None:
            print(f"  -> Aportes: '{headers[col_aportes]}'")
        if col_plan is not None:
            print(f"  -> Plan: '{headers[col_plan]}'")
        
        def get_value(row, col_idx):
            if col_idx is not None and col_idx < len(row):
                return row[col_idx] if row[col_idx] is not None else ""
            return ""
        
        def es_columna_texto(col_idx, sample_rows=10):
            """Verifica si una columna contiene texto largo (no números)"""
            if col_idx is None:
                return False
            
            # Verificar el nombre de la columna primero
            if col_idx < len(headers):
                col_name = str(headers[col_idx]).upper()
                # Si el nombre indica claramente que es texto, retornar True
                texto_keywords = ['COMENTARIO', 'APORTES', 'PLAN DE MEJORA', 'FORTALEZA', 'DEBILIDAD', 
                                 'QUE APORTES', 'ASPECTOS', 'OBJETIVOS', 'CONSIDERA', 'PROPUESTO']
                if any(keyword in col_name for keyword in texto_keywords):
                    # Verificar que no sea una columna numérica
                    if not any(num_keyword in col_name for num_keyword in ['PORCENTAJE', '%', 'PROMEDIO', 'CALIFICACION']):
                        return True
            
            # Verificar el contenido de las primeras filas
            texto_count = 0
            numero_count = 0
            
            for row_num in range(2, min(sample_rows + 2, sheet.max_row + 1)):
                row_data = [sheet.cell(row=row_num, column=col).value for col in range(1, sheet.max_column + 1)]
                value = get_value(row_data, col_idx)
                
                if value:
                    # Si es un string largo, es texto
                    if isinstance(value, str) and len(str(value).strip()) > 15:
                        texto_count += 1
                    # Si es un número, contar como número
                    elif isinstance(value, (int, float)):
                        numero_count += 1
            
            # Si hay más texto que números, es columna de texto
            return texto_count > numero_count and texto_count > 0
        
        # Procesar filas de datos
        for row_num in range(2, sheet.max_row + 1):
            row_data = [sheet.cell(row=row_num, column=col).value for col in range(1, sheet.max_column + 1)]
            
            nombre = get_value(row_data, col_nombre)
            if not nombre or str(nombre).strip() == "":
                continue
            
            # Extraer calificaciones específicas por tipo
            calificaciones = extraer_calificaciones_por_categoria(headers, row_data, tipo_evaluacion)
            
            # Calcular promedio inteligente
            promedio = 0
            
            # Prioridad 1: Usar promedio/porcentaje del Excel si existe y es válido
            if col_promedio is not None:
                valor_promedio = get_value(row_data, col_promedio)
                if valor_promedio and isinstance(valor_promedio, (int, float)):
                    try:
                        promedio = float(valor_promedio)
                        # Si está en escala 0-100, convertir a 1-5
                        if promedio > 5:
                            promedio = (promedio / 100) * 4 + 1
                        # Validar que esté en rango razonable
                        if 0 <= promedio <= 5:
                            pass  # Promedio válido del Excel
                        else:
                            promedio = 0  # Valor fuera de rango, calcular manualmente
                    except:
                        promedio = 0
            
            # Prioridad 2: Calcular de calificaciones si no hay promedio válido del Excel
            if promedio == 0:
                # Filtrar solo calificaciones válidas (entre 1 y 5)
                califs_validas = [v for v in calificaciones.values() if 1 <= v <= 5]
                if califs_validas:
                    promedio = sum(califs_validas) / len(califs_validas)
                else:
                    # Si no hay calificaciones específicas, buscar en todas las columnas numéricas
                    todas_califs = []
                    for i, val in enumerate(row_data):
                        if isinstance(val, (int, float)) and 1 <= val <= 5:
                            todas_califs.append(val)
                    
                    if todas_califs:
                        promedio = sum(todas_califs) / len(todas_califs)
            
            # Obtener textos (buscar en columnas de texto)
            comentario_jefe = ""
            if col_comentario is not None and es_columna_texto(col_comentario):
                val = get_value(row_data, col_comentario)
                # Solo usar si es texto y no un número
                if val and isinstance(val, str) and len(str(val).strip()) > 5:
                    comentario_jefe = str(val)
            
            aportes = ""
            if col_aportes is not None and es_columna_texto(col_aportes):
                val = get_value(row_data, col_aportes)
                # Solo usar si es texto y no un número
                if val and isinstance(val, str) and len(str(val).strip()) > 5:
                    aportes = str(val)
            
            plan_mejora = ""
            if col_plan is not None and es_columna_texto(col_plan):
                val = get_value(row_data, col_plan)
                # Solo usar si es texto y no un número
                if val and isinstance(val, str) and len(str(val).strip()) > 5:
                    plan_mejora = str(val)
            
            # Obtener periodo y fecha
            fecha_evaluacion = get_value(row_data, col_fecha)
            
            # Limpiar fecha: eliminar hora si existe y formatear
            if fecha_evaluacion:
                fecha_str = str(fecha_evaluacion)
                # Si es un objeto datetime de Excel
                if hasattr(fecha_evaluacion, 'strftime'):
                    fecha_str = fecha_evaluacion.strftime('%Y-%m-%d')
                else:
                    # Eliminar hora si existe (formato: "2024-01-15 00:00:00" -> "2024-01-15")
                    if ' ' in fecha_str:
                        fecha_str = fecha_str.split(' ')[0]
                fecha_evaluacion = fecha_str
            else:
                fecha_evaluacion = ""
            
            # Obtener periodo evaluado con validación
            periodo_evaluado = ""
            periodo_temp = ""
            
            # Intentar obtener el valor de la columna periodo
            if col_periodo is not None:
                periodo_temp = str(get_value(row_data, col_periodo)).strip()
            
            # VALIDACIÓN: Verificar que sea un periodo válido
            es_periodo_valido = False
            if periodo_temp and periodo_temp != "None" and periodo_temp != "":
                # Lista de palabras que indican que NO es un periodo
                palabras_invalidas = ['ANALISIS', 'ANÁLISIS', 'COMENTARIO', 'OBSERVACION', 
                                     'OBSERVACIÓN', 'FORTALEZA', 'DEBILIDAD', 'APORTES', 
                                     'PLAN', 'MEJORA', 'QUE', 'COMO', 'CONSIDERA', 'TRA',
                                     'BAJO', 'MEDIO', 'ALTO', 'RESPONSABILIDAD', 'GESTION',
                                     'GESTIÓN']
                
                periodo_upper = periodo_temp.upper()
                contiene_invalida = False
                
                for palabra in palabras_invalidas:
                    if palabra in periodo_upper:
                        contiene_invalida = True
                        print(f"[{nombre}] Periodo rechazado: '{periodo_temp}' (contiene '{palabra}')")
                        break
                
                # Validar longitud (un periodo no debería ser muy largo)
                if len(periodo_temp) > 25:
                    contiene_invalida = True
                    print(f"[{nombre}] Periodo rechazado: '{periodo_temp}' (muy largo: {len(periodo_temp)} caracteres)")
                
                # Si pasa todas las validaciones, es válido
                if not contiene_invalida:
                    es_periodo_valido = True
                    periodo_evaluado = periodo_temp
                    print(f"[{nombre}] Periodo aceptado del Excel: '{periodo_evaluado}'")
            
            # Si NO hay periodo válido, extraer año de la fecha
            if not es_periodo_valido or not periodo_evaluado:
                print(f"[{nombre}] Extrayendo año de fecha: '{fecha_evaluacion}'")
                
                try:
                    from datetime import datetime
                    import re
                    
                    año_extraido = ""
                    
                    # Método 1: Parsear fecha en formatos comunes
                    formatos = ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y', '%Y%m%d', '%d.%m.%Y']
                    for fmt in formatos:
                        try:
                            fecha_obj = datetime.strptime(fecha_evaluacion, fmt)
                            año_extraido = str(fecha_obj.year)
                            print(f"[{nombre}]   -> Año extraído (formato {fmt}): {año_extraido}")
                            break
                        except:
                            continue
                    
                    # Método 2: Buscar patrón de año (20XX o 19XX)
                    if not año_extraido:
                        year_match = re.search(r'(19|20)\d{2}', fecha_evaluacion)
                        if year_match:
                            año_extraido = year_match.group()
                            print(f"[{nombre}]   -> Año extraído (regex): {año_extraido}")
                    
                    # Método 3: Valor por defecto
                    if not año_extraido:
                        año_extraido = "2025"
                        print(f"[{nombre}]   -> Usando año por defecto: {año_extraido}")
                    
                    periodo_evaluado = año_extraido
                    
                except Exception as e:
                    print(f"[{nombre}]   -> Error: {e}, usando 2025")
                    periodo_evaluado = "2025"
            
            evaluacion = {
                'id': row_num - 1,
                'nombre': str(nombre).upper(),
                'cargo': str(get_value(row_data, col_cargo)).upper(),
                'area': str(get_value(row_data, col_area)).upper(),
                'jefe': str(get_value(row_data, col_jefe)).upper(),
                'fecha': fecha_evaluacion,
                'periodo': periodo_evaluado,
                'promedio': round(promedio, 1),
                'rendimiento': calcular_rendimiento(promedio),
                'comentario_jefe': comentario_jefe.upper(),
                'aportes': aportes.upper(),
                'plan_mejora': plan_mejora.upper(),
                'tipo_evaluacion': tipo_evaluacion,
                'calificaciones': calificaciones
            }
            
            evaluaciones_data.append(evaluacion)
        
        return evaluaciones_data
        
    except Exception as e:
        print(f"Error al procesar Excel: {e}")
        raise e

def generar_pdf_reportlab(evaluacion):
    """Genera PDF usando ReportLab replicando exactamente el diseño del HTML original"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch, 
                           leftMargin=0.5*inch, rightMargin=0.5*inch)
    
    # Estilos que replican el CSS original
    styles = getSampleStyleSheet()
    
    # Estilo para el título principal
    main_title_style = ParagraphStyle(
        'MainTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=6,
        alignment=1,  # Centrado
        fontName='Helvetica-Bold',
        textColor=colors.black
    )
    
    # Estilo para subtítulo
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=12,
        alignment=1,  # Centrado
        fontName='Helvetica',
        textColor=colors.black
    )
    
    # Estilo para títulos de sección
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=6,
        spaceBefore=12,
        fontName='Helvetica-Bold',
        textColor=colors.black,
        borderWidth=1,
        borderColor=colors.lightgrey,
        borderPadding=3
    )
    
    # Estilo para texto normal
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=3,
        fontName='Helvetica',
        textColor=colors.black
    )
    
    # Estilo para labels en negrita
    label_style = ParagraphStyle(
        'LabelStyle',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=3,
        fontName='Helvetica-Bold',
        textColor=colors.black
    )
    
    story = []
    
    # === HEADER CON LOGO Y CÓDIGOS ===
    # Intentar cargar el logo
    logo_path = os.path.join('static', 'logo.png')
    header_data = []
    
    if os.path.exists(logo_path):
        from reportlab.lib.utils import ImageReader
        try:
            # Crear tabla para el header con logo y códigos
            logo_img = ImageReader(logo_path)
            # Redimensionar logo a altura máxima de 45px (como en CSS)
            logo_height = 45
            logo_width = logo_height * (logo_img.getSize()[0] / logo_img.getSize()[1])
            
            header_data = [
                [
                    # Logo
                    logo_img,
                    # Códigos
                    Paragraph("<b>CODIGO:</b> FT-RH-042<br/><b>VERSION:</b> 1<br/><b>VIGENCIA:</b> 2026/02/03", 
                             ParagraphStyle('CodeStyle', parent=styles['Normal'], fontSize=10, alignment=2))
                ]
            ]
            
            header_table = Table(header_data, colWidths=[2*inch, 2*inch])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),   # Logo a la izquierda
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),  # Códigos a la derecha
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(header_table)
            
        except Exception as e:
            print(f"Error cargando logo: {e}")
            # Fallback sin logo
            header_data = [
                [
                    Paragraph("<b>LABORATORIOS NOVADERMA S.A.</b>", 
                             ParagraphStyle('CompanyName', parent=styles['Normal'], fontSize=20, fontName='Helvetica-Bold')),
                    Paragraph("<b>CODIGO:</b> FT-RH-042<br/><b>VERSION:</b> 1<br/><b>VIGENCIA:</b> 2026/02/03", 
                             ParagraphStyle('CodeStyle', parent=styles['Normal'], fontSize=10, alignment=2))
                ]
            ]
            header_table = Table(header_data, colWidths=[4*inch, 2*inch])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ]))
            story.append(header_table)
    else:
        # Sin logo disponible
        header_data = [
            [
                Paragraph("<b>LABORATORIOS NOVADERMA S.A.</b>", 
                         ParagraphStyle('CompanyName', parent=styles['Normal'], fontSize=20, fontName='Helvetica-Bold')),
                Paragraph("<b>CODIGO:</b> FT-RH-042<br/><b>VERSION:</b> 1<br/><b>VIGENCIA:</b> 2026/02/03", 
                         ParagraphStyle('CodeStyle', parent=styles['Normal'], fontSize=10, alignment=2))
            ]
        ]
        header_table = Table(header_data, colWidths=[4*inch, 2*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(header_table)
    
    # === TÍTULOS PRINCIPALES ===
    story.append(Paragraph("RESULTADOS EVALUACIÓN DE DESEMPEÑO LABORAL", main_title_style))
    story.append(Paragraph("Evaluación de desempeño y competencias del trabajador", subtitle_style))
    
    # === SECCIÓN 1: DATOS DE IDENTIFICACIÓN ===
    story.append(Paragraph("1. Datos de identificación del colaborador", section_title_style))
    
    datos_identificacion = [
        ["Empresa:", "LABORATORIOS NOVADERMA S.A."],
        ["Área / Proceso:", evaluacion.get('area', '').upper()],
        ["Cargo:", evaluacion.get('cargo', '').upper()],
        ["Nombre del colaborador:", evaluacion.get('nombre', '').upper()],
        ["Jefe inmediato:", evaluacion.get('jefe', '').upper()],
        ["Período evaluado:", evaluacion.get('periodo', '').upper()],
        ["Fecha de evaluación:", evaluacion.get('fecha', '').upper()]
    ]
    
    for label, value in datos_identificacion:
        story.append(Paragraph(f"<b>{label}</b> {value}", normal_style))
    
    story.append(Spacer(1, 12))
    
    # === SECCIÓN 2: RESUMEN DEL DESEMPEÑO ===
    story.append(Paragraph("2. Resumen del desempeño", section_title_style))
    
    # Crear tabla para promedio y rendimiento (como en el HTML original)
    promedio = evaluacion.get('promedio', 0)
    rendimiento = evaluacion.get('rendimiento', 'NO EVALUADO')
    
    resumen_data = [
        [f"PROMEDIO: {promedio:.1f}", f"RENDIMIENTO: {rendimiento.upper()}"]
    ]
    
    resumen_table = Table(resumen_data, colWidths=[2.5*inch, 2.5*inch])
    resumen_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(resumen_table)
    story.append(Spacer(1, 6))
    
    # Comentario del jefe
    comentario = evaluacion.get('comentario_jefe', 'Sin comentarios').upper()
    story.append(Paragraph(f"<b>Comentario del jefe inmediato:</b><br/>{comentario}", normal_style))
    
    story.append(Spacer(1, 12))
    
    # === SECCIÓN 3: DESEMPEÑO EN EL CARGO ===
    story.append(Paragraph("3. Desempeño en el cargo (1 a 5)", section_title_style))
    
    calificaciones = evaluacion.get('calificaciones', {})
    tipo_evaluacion = evaluacion.get('tipo_evaluacion', 'OPERATIVO')
    
    # Crear tabla de calificaciones
    cal_data = []
    
    # Calificaciones base adaptadas por tipo
    if calificaciones.get('organizacion', 0) > 0:
        if tipo_evaluacion == 'COMERCIAL':
            label = "Planificación y organización del trabajo comercial:"
        elif tipo_evaluacion == 'DIRECTIVOS':
            label = "Organización y planificación estratégica:"
        elif tipo_evaluacion == 'ADMINISTRATIVA':
            label = "Organización de tareas administrativas:"
        else:
            label = "Organización del trabajo y cumplimiento de tiempos:"
        cal_data.append([label, f"{calificaciones['organizacion']} / 5"])
    
    if calificaciones.get('cumple_resultados', 0) > 0:
        if tipo_evaluacion == 'COMERCIAL':
            label = "Cumplimiento de objetivos comerciales y cuotas:"
        elif tipo_evaluacion == 'DIRECTIVOS':
            label = "Cumplimiento de resultados del área dirigida:"
        elif tipo_evaluacion == 'ADMINISTRATIVA':
            label = "Cumplimiento de tareas y procedimientos:"
        else:
            label = "Cumple con los resultados esperados de su función:"
        cal_data.append([label, f"{calificaciones['cumple_resultados']} / 5"])
    
    if calificaciones.get('aplica_capacitacion', 0) > 0:
        cal_data.append(["Aplica conceptos de capacitaciones y entrenamientos:", f"{calificaciones['aplica_capacitacion']} / 5"])
    
    # Calificaciones específicas por tipo
    if tipo_evaluacion == 'COMERCIAL':
        if calificaciones.get('ventas', 0) > 0:
            cal_data.append(["Cumplimiento de cuotas y panel médico:", f"{calificaciones['ventas']} / 5"])
        if calificaciones.get('clientes', 0) > 0:
            cal_data.append(["Atención y relación con clientes/médicos:", f"{calificaciones['clientes']} / 5"])
    
    elif tipo_evaluacion == 'DIRECTIVOS':
        if calificaciones.get('liderazgo', 0) > 0:
            cal_data.append(["Liderazgo y dirección de equipos:", f"{calificaciones['liderazgo']} / 5"])
        if calificaciones.get('gestion', 0) > 0:
            cal_data.append(["Gestión eficiente de recursos:", f"{calificaciones['gestion']} / 5"])
        if calificaciones.get('evaluacion_equipo', 0) > 0:
            cal_data.append(["Evaluación y análisis de resultados del equipo:", f"{calificaciones['evaluacion_equipo']} / 5"])
    
    elif tipo_evaluacion == 'ADMINISTRATIVA':
        if calificaciones.get('analisis', 0) > 0:
            cal_data.append(["Análisis y solución de problemas:", f"{calificaciones['analisis']} / 5"])
        if calificaciones.get('informes', 0) > 0:
            cal_data.append(["Entrega de informes y tareas:", f"{calificaciones['informes']} / 5"])
        if calificaciones.get('planificacion', 0) > 0:
            cal_data.append(["Planificación y cumplimiento del plan de trabajo:", f"{calificaciones['planificacion']} / 5"])
    
    # Agregar más calificaciones comunes
    for key, label in [
        ('uso_equipos', 'Uso adecuado de equipos y recursos:'),
        ('cumple_politicas', 'Cumple con políticas, procedimientos, normas y horarios:'),
        ('conoce_calidad', 'Conoce y aplica la política de calidad:'),
        ('propone_mejoras', 'Propone alternativas de mejora e ideas constructivas:'),
        ('relaciones', 'Relaciones interpersonales y cordialidad:'),
        ('trabajo_equipo', 'Trabajo en equipo y colaboración:'),
        ('actitud_servicio', 'Actitud de servicio:')
    ]:
        if calificaciones.get(key, 0) > 0:
            cal_data.append([label, f"{calificaciones[key]} / 5"])
    
    if cal_data:
        cal_table = Table(cal_data, colWidths=[4.5*inch, 1*inch])
        cal_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ]))
        story.append(cal_table)
    
    story.append(Spacer(1, 12))
    
    # === SECCIÓN 6: APORTES ===
    story.append(Paragraph("6. Aportes", section_title_style))
    aportes = evaluacion.get('aportes', 'Sin aportes registrados').upper()
    story.append(Paragraph(aportes, normal_style))
    
    story.append(Spacer(1, 12))
    
    # === SECCIÓN 7: PLAN DE MEJORA ===
    story.append(Paragraph("7. Plan de mejora", section_title_style))
    plan_mejora = evaluacion.get('plan_mejora', 'Sin plan de mejora registrado').upper()
    story.append(Paragraph(f"<b>Acción de mejora:</b><br/>{plan_mejora}", normal_style))
    
    # === FOOTER ===
    story.append(Spacer(1, 20))
    footer_text = """Este formato se utiliza como registro de evaluación de desempeño y competencias del trabajador, 
    en coherencia con las políticas internas de LABORATORIOS NOVADERMA S.A. y con los lineamientos 
    del Sistema de Gestión de la Calidad."""
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.black,
        borderWidth=1,
        borderColor=colors.lightgrey,
        borderPadding=5,
        spaceBefore=10
    )
    
    story.append(Paragraph(footer_text, footer_style))
    
    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/test-pdf')
def test_pdf():
    """Endpoint de prueba para verificar que la generación de PDF funciona"""
    try:
        # Intentar primero con WeasyPrint
        try:
            html_simple = """
            <!DOCTYPE html>
            <html>
            <head><meta charset="UTF-8"></head>
            <body>
                <h1>Prueba de PDF con WeasyPrint</h1>
                <p>Si ves esto, WeasyPrint funciona correctamente.</p>
            </body>
            </html>
            """
            pdf_bytes = HTML(string=html_simple).write_pdf()
            pdf_buffer = BytesIO(pdf_bytes)
            pdf_buffer.seek(0)
            
            return send_file(pdf_buffer, 
                            as_attachment=True,
                            download_name='test_weasyprint.pdf',
                            mimetype='application/pdf')
                            
        except Exception as weasy_error:
            print(f"WeasyPrint falló: {weasy_error}")
            
            # Usar ReportLab como alternativa
            if REPORTLAB_AVAILABLE:
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4)
                
                styles = getSampleStyleSheet()
                story = []
                
                story.append(Paragraph("Prueba de PDF con ReportLab", styles['Title']))
                story.append(Spacer(1, 12))
                story.append(Paragraph("Si ves esto, ReportLab funciona correctamente.", styles['Normal']))
                story.append(Spacer(1, 12))
                story.append(Paragraph("WeasyPrint no está disponible, usando ReportLab como alternativa.", styles['Normal']))
                
                doc.build(story)
                buffer.seek(0)
                
                return send_file(buffer, 
                                as_attachment=True,
                                download_name='test_reportlab.pdf',
                                mimetype='application/pdf')
            else:
                return f"Error: Ni WeasyPrint ni ReportLab están disponibles.<br><br>Error de WeasyPrint: {str(weasy_error)}", 500
                
    except Exception as e:
        import traceback
        return f"Error general: {str(e)}<br><br><pre>{traceback.format_exc()}</pre>", 500

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        if file and file.filename.lower().endswith(('.xlsx', '.xls')):
            # Guardar archivo temporalmente
            filename = file.filename
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            
            # Procesar Excel
            evaluaciones = procesar_excel(filepath)
            
            # Limpiar archivo temporal
            try:
                os.remove(filepath)
            except:
                pass
            
            return jsonify({
                'success': True,
                'evaluaciones': evaluaciones,
                'total': len(evaluaciones)
            })
        else:
            return jsonify({'error': 'Formato de archivo no válido. Use .xlsx o .xls'}), 400
            
    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500

@app.route('/generar-pdf/<int:eval_id>', methods=['POST'])
def generar_pdf(eval_id):
    try:
        data = request.json
        evaluacion = data.get('evaluacion')
        
        if not evaluacion:
            return jsonify({'error': 'No se recibieron datos'}), 400
        
        # Obtener logo
        try:
            logo_base64 = get_logo_base64()
        except Exception as e:
            print(f"Advertencia: No se pudo cargar el logo: {e}")
            logo_base64 = ''
        
        # Renderizar HTML
        html_content = render_template('reporte.html', 
                                      evaluacion=evaluacion,
                                      logo_base64=logo_base64)
        
        # Intentar primero con WeasyPrint (método preferido)
        try:
            print("Intentando generar PDF con WeasyPrint...")
            
            # Configurar WeasyPrint
            from weasyprint import HTML
            
            # Generar PDF con WeasyPrint sin CSS extra
            html_doc = HTML(string=html_content, base_url='.')
            pdf_bytes = html_doc.write_pdf()
            pdf_buffer = BytesIO(pdf_bytes)
            pdf_buffer.seek(0)
            
            print(f"PDF generado exitosamente con WeasyPrint. Tamaño: {len(pdf_bytes)} bytes")
            
        except Exception as weasy_error:
            print(f"WeasyPrint falló: {weasy_error}")
            import traceback
            print(f"Traceback completo:\n{traceback.format_exc()}")
            
            # Intentar con pdfkit como segunda opción
            try:
                if PDFKIT_AVAILABLE:
                    print("Intentando con pdfkit...")
                    # Configuración para pdfkit
                    options = {
                        'page-size': 'A4',
                        'margin-top': '0.75in',
                        'margin-right': '0.75in',
                        'margin-bottom': '0.75in',
                        'margin-left': '0.75in',
                        'encoding': "UTF-8",
                        'no-outline': None,
                        'enable-local-file-access': None
                    }
                    
                    pdf_bytes = pdfkit.from_string(html_content, False, options=options)
                    pdf_buffer = BytesIO(pdf_bytes)
                    pdf_buffer.seek(0)
                    
                    print(f"PDF generado exitosamente con pdfkit. Tamaño: {len(pdf_bytes)} bytes")
                    
                else:
                    raise Exception("pdfkit no disponible")
                    
            except Exception as pdfkit_error:
                print(f"pdfkit también falló: {pdfkit_error}")
                
                # Usar ReportLab como última alternativa
                if REPORTLAB_AVAILABLE:
                    print("Usando ReportLab como última alternativa...")
                    pdf_buffer = generar_pdf_reportlab(evaluacion)
                    print("PDF generado con ReportLab (funcionalidad limitada)")
                else:
                    return jsonify({'error': f'Error al generar PDF: WeasyPrint: {str(weasy_error)}, pdfkit: {str(pdfkit_error)}'}), 500
        
        # Nombre del archivo (sanitizar caracteres especiales)
        nombre_sanitizado = ''.join(c if c.isalnum() or c in (' ', '_') else '_' for c in evaluacion['nombre'])
        pdf_filename = f"evaluacion_{nombre_sanitizado.replace(' ', '_')}_{eval_id}.pdf"
        
        return send_file(pdf_buffer, 
                        as_attachment=True,
                        download_name=pdf_filename,
                        mimetype='application/pdf')
                        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error general al generar PDF:\n{error_detail}")
        return jsonify({'error': f'Error al generar PDF: {str(e)}'}), 500

if __name__ == '__main__':
    # En desarrollo
    app.run(debug=True, host='0.0.0.0', port=5000)